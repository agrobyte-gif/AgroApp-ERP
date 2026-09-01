"""Saca a una planilla los productos que no tienen peso, con una propuesta.

    odoo-bin shell -c config/odoo.conf -d agrogood_dev --no-http < tools/exportar_pesos.py

Escribe `C:/dev/Pesos por completar.xlsx`, fuera del repositorio.

Por que hace falta: la ruta suma el peso de lo que lleva para avisar si el
camion va sobrecargado. En los productos de peso variable el peso ES la
cantidad y sale solo; en los que se venden por unidad -una caja, un atado, una
malla- hay que decirlo, y ninguno lo tiene. Con todos a cero, el aviso de
sobrecarga no salta nunca y el camion se carga a ojo.

La propuesta se calcula por FORMATO y no por categoria. Se probaron las dos:
por categoria, en VERDURAS conviven una acelga y una caja, y en FRUTAS una pina
pesa doce veces menos que una caja de manzanas. El formato -caja, saco, malla,
atado- si dice cuanto pesa la cosa, que es de lo que se trata.

Nadie tiene que teclear 73 pesos: la columna viene rellena y solo hay que
corregir lo que no cuadre. Un peso aproximado sirve; cero no sirve.
"""

import os
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SALIDA = r"C:\dev\Pesos por completar.xlsx"

# Kilos por unidad vendida. Son valores de partida de distribucion
# hortofruticola, para que la planilla no salga vacia: Bodega los corrige en
# los que no cuadren, que siempre son unos pocos.
POR_FORMATO = {
    'Caja': 15.0,
    'Saco': 25.0,
    'Malla': 5.0,
    'Bandeja': 1.0,
    'Bolsa': 1.0,
    'Atado': 0.5,
    'Granel': 1.0,
}
SUELTO = 0.4          # una pieza o un atado sin formato declarado

# Cuando el formato no esta asignado, el nombre suele delatarlo.
PISTAS = [
    (r"\bSACO\b", 'Saco'),
    (r"\bCAJA\b", 'Caja'),
    (r"\bMALLA\b", 'Malla'),
    (r"\bBANDEJA\b", 'Bandeja'),
    (r"\bBOLSA\b", 'Bolsa'),
    (r"\b(ATADO|PQTE|PAQUETE)\b", 'Atado'),
]


def propuesta(producto, formato):
    """Kilos propuestos para una unidad de venta.

    Lo que se vende por LITROS se resuelve solo: un litro de liquido pesa
    aproximadamente un kilo. Eso es fisica y no una adivinanza, asi que no hay
    que preguntarlo.

    El resto sale del formato. Y el valor de la pieza suelta -0,4 kg- es un
    promedio de atado y manojo que se queda MUY corto en lo voluminoso: una
    sandia pesa siete kilos y una pina dos. Se deja asi a proposito, en lugar
    de inventar una lista de excepciones que estaria incompleta el primer dia:
    la planilla existe justamente para corregir esos, y verlos disparatados en
    la columna hace que se corrijan.
    """
    if (producto.uom_id.name or "").upper().startswith("L"):
        return 1.0
    return POR_FORMATO.get(formato, SUELTO)


def formato_de(producto):
    """El formato declarado; si no hay, el que delata el nombre."""
    if producto.agrogood_format_id:
        return producto.agrogood_format_id.name, "declarado"
    nombre = (producto.name or "").upper()
    for patron, formato in PISTAS:
        if re.search(patron, nombre):
            return formato, "deducido del nombre"
    return "", "suelto"


Producto = env['product.template']
pendientes = Producto.search([
    ('is_storable', '=', True),
    ('agrogood_is_variable_weight', '=', False),
]).filtered(lambda p: not p.weight)

print("=" * 74)
print("PESOS POR COMPLETAR")
print("=" * 74)
print("Productos de peso fijo sin peso: %d" % len(pendientes))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pesos"

CABECERAS = ["Codigo", "Producto", "Categoria", "Se vende en",
             "Formato", "De donde sale", "PESO KG (corregir)"]
ws.append(CABECERAS)
for i, celda in enumerate(ws[1], start=1):
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="1C874F")
    celda.alignment = Alignment(horizontal="center")

resumen = {}
for p in sorted(pendientes, key=lambda x: (x.categ_id.complete_name, x.name)):
    formato, origen = formato_de(p)
    peso = propuesta(p, formato)
    resumen[formato or "(suelto)"] = resumen.get(formato or "(suelto)", 0) + 1
    ws.append([p.default_code or "", p.name,
               p.categ_id.complete_name.replace("Agrogood / ", ""),
               p.uom_id.name, formato, origen, peso])

anchos = [12, 38, 20, 12, 12, 20, 18]
for i, ancho in enumerate(anchos, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
ws.freeze_panes = "A2"

# La columna que hay que tocar, marcada. Una planilla de siete columnas donde
# solo se edita una tiene que decir cual sin que nadie lo explique.
for fila in ws.iter_rows(min_row=2, min_col=7, max_col=7):
    for celda in fila:
        celda.fill = PatternFill("solid", fgColor="FFF3CD")
        celda.number_format = '0.00'

# Una segunda hoja con los valores de partida, para que se vean y se discutan.
ws2 = wb.create_sheet("De donde salen")
ws2.append(["Formato", "Kilos propuestos por unidad"])
ws2["A1"].font = ws2["B1"].font = Font(bold=True)
for formato, kilos in sorted(POR_FORMATO.items()):
    ws2.append([formato, kilos])
ws2.append(["(sin formato: pieza o atado suelto)", SUELTO])
ws2.column_dimensions['A'].width = 38
ws2.column_dimensions['B'].width = 26

wb.save(SALIDA)

print()
print("%-22s %s" % ("FORMATO", "PRODUCTOS"))
for formato, n in sorted(resumen.items(), key=lambda x: -x[1]):
    kilos = POR_FORMATO.get(formato, SUELTO)
    print("%-22s %4d   se propone %.2f kg" % (formato, n, kilos))
print()
print("Planilla escrita en: %s" % SALIDA)
print("Existe: %s" % os.path.exists(SALIDA))
print()
print("Solo hay que revisar la ultima columna, que viene rellena. Corregir lo")
print("que no cuadre y volver a guardarla; despues:")
print("   AGROGOOD_PESOS=escribir odoo-bin shell ... < tools/importar_pesos.py")
