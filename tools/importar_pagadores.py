"""Ensena a Agroapp quien es cada pagador, desde la planilla ya llena.

    odoo-bin shell -c config/odoo.conf -d agrogood_dev --no-http < tools/importar_pagadores.py

Lee C:\\dev\\Pagadores por identificar.xlsx -sus dos hojas- guarda cada RUT y
cada nombre del banco como identidad de pago del cliente que se escribio al
lado, y vuelve a cruzar los abonos.

Tres decisiones que conviene tener a la vista:

1. **No se crean clientes.** Si el nombre escrito no existe en Odoo, la fila se
   informa y se deja quieta. Crear la ficha aqui llenaria la lista de clientes
   de variantes tipograficas del mismo negocio, que despues hay que fusionar a
   mano y mientras tanto parten la cuenta corriente en dos.

2. **Una identidad que ya es de otro cliente no se pisa: se marca compartida.**
   Es lo que hace `aprender()`, y es lo que pasa de verdad -la sociedad que
   paga por dos locales-. A partir de ahi esa identidad espera a una persona
   en vez de asignar sola.

3. **Se aprende, y recien despues se cruza.** Cruzar primero no serviria de
   nada: lo que hace que el cruce encuentre algo nuevo es justamente lo que se
   acaba de ensenar.

Se puede correr las veces que haga falta: lo que ya estaba se cuenta aparte y
no se vuelve a escribir.
"""

import os

try:
    import openpyxl
except ImportError:
    raise SystemExit("Falta openpyxl: .venv\\Scripts\\pip install openpyxl")

from odoo.addons.agrogood_crm_reactivation.models.agrogood_payer import (
    normalizar_alias, normalizar_rut)

ORIGEN = r"C:\dev\Pagadores por identificar.xlsx"
if not os.path.exists(ORIGEN):
    raise SystemExit("No esta %s. Primero: tools/exportar_pagadores.py" % ORIGEN)

P = env['res.partner']
Pagador = env['agrogood.payer']
M = env['agrogood.bank.movement']


def compacto(t):
    return " ".join(str(t or "").split()).upper()


# Cuando en la columna del cliente se escribe una de estas, no es que falte el
# nombre: es que ese abono no es un cobro -un deposito, un cheque, plata que se
# mueve entre cuentas propias-. En vez de dejarlo esperando un cliente que no
# existe, se descarta con el motivo a la vista. Es la respuesta que Victor da
# mas seguido despues de "quien es": "ese no es cliente".
NO_ES_CLIENTE = {
    "NO ES CLIENTE", "NOES CLIENTE", "NO CLIENTE", "NO", "NO APLICA",
    "CUENTA PROPIA", "PROPIA", "DEPOSITO", "CHEQUE", "VALE VISTA",
    "DEVOLUCION", "PRESTAMO", "REVERSO",
}


# El indice se arma una vez: buscar cliente por cliente con ilike sobre 600
# filas es lento y ademas encuentra parecidos, que es justo lo que no se
# quiere aqui.
indice = {}
for c in P.search([('customer_rank', '>', 0)]):
    indice.setdefault(compacto(c.name), c)

libro = openpyxl.load_workbook(ORIGEN, data_only=True)
resumen = []
sin_ficha = {}
descartados = {'n': 0}


def leer_hoja(nombre_hoja, tipo):
    """Aprende una hoja. `tipo` es 'rut' o 'alias'."""
    if nombre_hoja not in libro.sheetnames:
        print("  (no hay hoja '%s' en la planilla)" % nombre_hoja)
        return
    hoja = libro[nombre_hoja]
    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        return
    cab = [str(c or "").strip().upper() for c in filas[0]]

    def cual(*claves):
        for i, c in enumerate(cab):
            if any(k in c for k in claves):
                return i
        raise SystemExit("Falta la columna %s en la hoja %s"
                         % (claves, nombre_hoja))

    icl, ino = 0, cual("CLIENTE")
    nuevas, ya, vacias = 0, 0, 0

    for f in filas[1:]:
        if ino >= len(f):
            continue
        valor = (normalizar_rut(f[icl]) if tipo == 'rut'
                 else normalizar_alias(f[icl]))
        nombre = str(f[ino] or "").strip()
        if not valor:
            continue
        if not nombre:
            vacias += 1
            continue
        if compacto(nombre) in NO_ES_CLIENTE:
            if tipo == 'rut':
                movs = M.search([('state', 'in', ('unknown', 'doubtful')),
                                 ('payer_rut', '=', valor)])
            else:
                movs = M.search([('state', 'in', ('unknown', 'doubtful')),
                                 ('payer_alias', '!=', False)]).filtered(
                    lambda m: normalizar_alias(m.payer_alias) == valor)
            if movs:
                movs.write({'state': 'discarded', 'partner_id': False,
                            'match_reason': "No es un cobro: %s" % nombre})
                descartados['n'] += len(movs)
            continue
        cliente = indice.get(compacto(nombre))
        if not cliente:
            sin_ficha[nombre] = sin_ficha.get(nombre, 0) + 1
            continue
        existe = Pagador.search(
            [('kind', '=', tipo), ('value', '=', valor)], limit=1)
        if existe and existe.partner_id == cliente:
            ya += 1
            continue
        Pagador.aprender(cliente, **{tipo: valor})
        nuevas += 1

    resumen.append((nombre_hoja, nuevas, ya, vacias))


print("=" * 74)
print("IDENTIDADES DE PAGO")
print("=" * 74)
leer_hoja("Nombres del banco", 'alias')
leer_hoja("Pagadores por RUT", 'rut')

for hoja, nuevas, ya, vacias in resumen:
    print()
    print("  %s" % hoja)
    print("      aprendidas ahora      : %d" % nuevas)
    print("      ya estaban            : %d" % ya)
    print("      sin nombre en la hoja : %d  (se dejan para despues)" % vacias)

if descartados['n']:
    print()
    print("  Abonos descartados por 'no es cliente': %d" % descartados['n'])

if sin_ficha:
    print()
    print("  NOMBRES QUE NO EXISTEN COMO CLIENTE: %d" % len(sin_ficha))
    for n in sorted(sin_ficha)[:15]:
        print("      %s" % n)
    if len(sin_ficha) > 15:
        print("      ... y %d mas" % (len(sin_ficha) - 15))
    print()
    print("  O se corrige la escritura en la planilla, o se crea la ficha")
    print("  antes. Aqui no se crean: cada variante tipografica del mismo")
    print("  negocio parte su cuenta corriente en dos.")

# Ahora si: se vuelve a mirar lo que estaba sin nombre.
antes = M.search_count([('state', '=', 'identified')])
pendientes = M.search([('state', 'in', ('unknown', 'doubtful'))])
print()
print("Volviendo a cruzar %d abonos..." % len(pendientes))
pendientes._cruzar()
env.cr.flush()


def plata(x):
    return "{:,.0f}".format(x).replace(",", ".")


print()
for est, etiqueta in (('identified', 'reconocidos'), ('unknown', 'sin nombre'),
                      ('doubtful', 'dudosos'), ('discarded', 'no son cobros')):
    r = M.search([('state', '=', est), ('amount', '>', 0)])
    print("  %-14s %5d   %16s"
          % (etiqueta, len(r), plata(sum(r.mapped('amount')))))

cobros = M.search([('state', 'in', ('identified', 'unknown', 'doubtful')),
                   ('amount', '>', 0)])
ident = M.search([('state', '=', 'identified'), ('amount', '>', 0)])
total = sum(cobros.mapped('amount')) or 1.0
print()
print("  reconocidos antes: %d   ahora: %d   (+%d)"
      % (antes, len(ident), len(ident) - antes))
print("  De la plata que si es cobro de cliente, se reconoce el %.0f%%"
      % (100.0 * sum(ident.mapped('amount')) / total))
print()
print("Nada esta guardado todavia. Para dejarlo: env.cr.commit()")
