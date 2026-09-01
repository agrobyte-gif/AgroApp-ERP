"""Carga los pesos desde la planilla completada.

    odoo-bin shell -c config/odoo.conf -d agrogood_dev --no-http < tools/importar_pesos.py

Por defecto SOLO INFORMA. Para escribir: AGROGOOD_PESOS=escribir

Lee `C:/dev/Pesos por completar.xlsx`, la que genera `exportar_pesos.py`. Se
busca el producto por su CODIGO y, si no lo tiene, por el nombre exacto. No se
adivina por parecido: un peso puesto en el producto equivocado no da error, da
un camion mal calculado que nadie va a revisar.

Un peso aproximado sirve; cero no sirve. Con todos a cero, el aviso de
sobrecarga no salta nunca y el camion se carga a ojo.
"""

import os

import openpyxl

RUTA = r"C:\dev\Pesos por completar.xlsx"
ESCRIBIR = os.environ.get("AGROGOOD_PESOS") == "escribir"

# Un peso disparatado casi siempre es un error de tecleo o una unidad
# equivocada -gramos en vez de kilos-. Se avisa en vez de aceptarlo: un saco de
# 25 kg escrito como 25.000 convierte cualquier ruta en sobrecargada y el aviso
# deja de creerse.
MAXIMO_RAZONABLE = 60.0

print("=" * 74)
print("CARGA DE PESOS" + ("  [ESCRIBIENDO]" if ESCRIBIR else "  [SOLO INFORME]"))
print("=" * 74)

if not os.path.exists(RUTA):
    print("No existe %s" % RUTA)
    print("Generala primero con tools/exportar_pesos.py")
else:
    ws = openpyxl.load_workbook(RUTA, data_only=True)["Pesos"]
    Producto = env['product.template']

    aplicables, sin_encontrar, sin_peso, sospechosos = [], [], [], []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila or not (fila[0] or fila[1]):
            continue
        codigo = str(fila[0] or "").strip()
        nombre = str(fila[1] or "").strip()
        try:
            peso = float(fila[6]) if fila[6] not in (None, "") else 0.0
        except (TypeError, ValueError):
            peso = 0.0

        producto = Producto.search([('default_code', '=', codigo)], limit=1) \
            if codigo else Producto.browse()
        if not producto:
            producto = Producto.search([('name', '=', nombre)], limit=1)
        if not producto:
            sin_encontrar.append(nombre or codigo)
            continue
        if peso <= 0:
            sin_peso.append(producto.display_name)
            continue
        if peso > MAXIMO_RAZONABLE:
            sospechosos.append((producto, peso))
            continue
        aplicables.append((producto, peso))

    print("\nFilas leidas de la planilla: %d" % (ws.max_row - 1))
    print("  con peso para aplicar   : %d" % len(aplicables))
    print("  sin peso (se dejan)     : %d" % len(sin_peso))
    print("  no se encontro el producto: %d" % len(sin_encontrar))
    print("  peso sospechoso (>%.0f kg): %d" % (MAXIMO_RAZONABLE, len(sospechosos)))

    if sospechosos:
        print("\n" + "-" * 74)
        print("NO SE APLICAN: un peso asi suele ser gramos escritos como kilos")
        for p, peso in sospechosos[:10]:
            print("  %-40s %.2f kg" % (p.display_name[:40], peso))

    if sin_encontrar:
        print("\n" + "-" * 74)
        print("NO SE ENCONTRARON (%d)" % len(sin_encontrar))
        for n in sin_encontrar[:10]:
            print("  %s" % n[:60])

    if aplicables:
        print("\n" + "-" * 74)
        print("SE APLICAN (muestra)")
        for p, peso in aplicables[:12]:
            print("  %-42s %8.2f kg" % (p.display_name[:42], peso))
        if len(aplicables) > 12:
            print("  ... y %d mas" % (len(aplicables) - 12))
        total = sum(peso for _, peso in aplicables)
        print("\n  peso medio propuesto: %.2f kg" % (total / len(aplicables)))

    if not ESCRIBIR:
        print("\n" + "=" * 74)
        print("SOLO INFORME. Nada se ha modificado.")
        print("Para aplicar: AGROGOOD_PESOS=escribir")
        print("=" * 74)
    else:
        for producto, peso in aplicables:
            producto.weight = peso
        env.cr.commit()
        quedan = Producto.search_count([
            ('is_storable', '=', True),
            ('agrogood_is_variable_weight', '=', False),
            ('weight', '=', 0)])
        print("\n" + "=" * 74)
        print("APLICADOS: %d pesos" % len(aplicables))
        print("Productos de peso fijo que siguen sin peso: %d" % quedan)
        print("=" * 74)
        print("\nEl aviso de sobrecarga del camion ya tiene con que calcular.")
