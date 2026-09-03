"""Saca la lista de quienes pagan y todavia no tienen nombre.

    odoo-bin shell -c config/odoo.conf -d agrogood_dev --no-http < tools/exportar_pagadores.py

Deja C:\\dev\\Pagadores por identificar.xlsx

Van DOS hojas, porque el banco identifica al pagador de dos maneras distintas
y cada una cuesta un trabajo distinto:

 * **Por nombre corto** (956 abonos, 2.001 millones). Santander no publica el
   RUT en estos: solo BAR CALLEJON, LOCO JOE, HOTEL DIEGO. Son 139 nombres y
   **diez de ellos son el 94% de esa plata**. Es la hoja que mas rinde por
   fila llenada, y por eso va primero.
 * **Por RUT** (7.885 abonos, 1.500 millones). Son 454 RUT, pero treinta de
   ellos son el 71% de la plata.

Las dos van ordenadas por monto con el % acumulado al lado, para poder parar
donde deje de valer la pena en vez de tener que llegar al final.

La columna que se llena es UNA: el nombre del cliente.
"""

import os
from collections import defaultdict

try:
    import xlsxwriter
except ImportError:
    raise SystemExit("Falta xlsxwriter: .venv\\Scripts\\pip install xlsxwriter")

DESTINO = r"C:\dev\Pagadores por identificar.xlsx"

M = env['agrogood.bank.movement']
sin_nombre = M.search([('state', '=', 'unknown'), ('amount', '>', 0),
                       ('payer_rut', '!=', False)])

# Se agrupa por RUT: el banco escribe el mismo pagador de diez maneras y lo que
# se identifica es la persona, no cada linea.
grupos = defaultdict(lambda: {'monto': 0.0, 'veces': 0, 'alias': set(),
                              'primera': None, 'ultima': None})
for m in sin_nombre:
    g = grupos[m.payer_rut]
    g['monto'] += m.amount
    g['veces'] += 1
    texto = (m.payer_alias or m.payer_name or m.description or "").strip()
    if texto:
        g['alias'].add(texto[:40])
    if not g['primera'] or m.date < g['primera']:
        g['primera'] = m.date
    if not g['ultima'] or m.date > g['ultima']:
        g['ultima'] = m.date

orden = sorted(grupos.items(), key=lambda x: -x[1]['monto'])

# Y los que el banco no identifica con un RUT: solo traen el nombre corto. En
# Santander son casi 1.000 abonos por 2.000 millones, mas plata que todos los
# que si traen RUT juntos.
solo_alias = M.search([('state', '=', 'unknown'), ('amount', '>', 0),
                       ('payer_rut', '=', False), ('payer_alias', '!=', False)])
por_nombre = defaultdict(lambda: {'monto': 0.0, 'veces': 0, 'alias': set(),
                                  'primera': None, 'ultima': None})
for m in solo_alias:
    clave = " ".join((m.payer_alias or "").split()).upper()
    if not clave:
        continue
    g = por_nombre[clave]
    g['monto'] += m.amount
    g['veces'] += 1
    if (m.description or "").strip():
        g['alias'].add(m.description.strip()[:40])
    if not g['primera'] or m.date < g['primera']:
        g['primera'] = m.date
    if not g['ultima'] or m.date > g['ultima']:
        g['ultima'] = m.date
orden_alias = sorted(por_nombre.items(), key=lambda x: -x[1]['monto'])

# Si el RUT ya esta en una ficha, no hay nada que preguntar: se dice cual es.
P = env['res.partner']
def ficha(rut):
    c = P.search([('vat', 'in', (rut, 'CL' + rut))], limit=1)
    return c.name if c else ""

# Y si no esta en una ficha, puede estar en la planilla de RUT que Victor ya
# tenia. Esa planilla no es una ficha -son empresas, no clientes de Odoo- pero
# sirve para no volver a preguntar 97 nombres que ya estaban escritos.
PLANILLA = r"C:\dev\DATOS CLIENTES RUT.xlsx"
conocidos = {}
if os.path.exists(PLANILLA):
    import openpyxl
    from odoo.addons.agrogood_crm_reactivation.models.agrogood_payer import (
        normalizar_rut)
    hoja = openpyxl.load_workbook(PLANILLA, data_only=True).active
    filas = list(hoja.iter_rows(values_only=True))
    cab = [str(c or "").strip().upper() for c in filas[0]]
    def cual(*claves):
        for i, c in enumerate(cab):
            if any(k in c for k in claves):
                return i
        return None
    iru = cual("RUT")
    ino = cual("EMPRESA", "NOMBRE", "CLIENTE", "RAZON")
    for f in filas[1:]:
        if iru is None or iru >= len(f):
            continue
        r = normalizar_rut(f[iru])
        if r and ino is not None and ino < len(f):
            conocidos[r] = str(f[ino] or "").strip()

if os.path.exists(DESTINO):
    raise SystemExit("Ya existe %s. Se renombra o se borra antes." % DESTINO)

libro = xlsxwriter.Workbook(DESTINO)
titulo = libro.add_format({'bold': True, 'bg_color': '#1F5C3A',
                           'font_color': 'white', 'border': 1})
pesos = libro.add_format({'num_format': '#,##0'})
llenar = libro.add_format({'bg_color': '#FFF3C4', 'border': 1})
gris = libro.add_format({'font_color': '#888888'})


def escribir(nombre_hoja, encabezado, filas, proponer):
    """Una hoja: la clave, cuanto movio, y el hueco para el nombre.

    `proponer` devuelve (nombre, de donde salio). Se propone en vez de dejarlo
    en blanco porque decenas de esos nombres ya estaban escritos en una
    planilla que Victor tenia, y volver a preguntarlos seria hacerle el mismo
    trabajo dos veces.
    """
    h = libro.add_worksheet(nombre_hoja)
    cols = [encabezado, "Como aparece en el banco", "Cuanto pago", "Abonos",
            "Desde", "Hasta", "% acumulado", "CLIENTE (llenar aqui)",
            "De donde salio"]
    for i, c in enumerate(cols):
        h.write(0, i, c, titulo)
    h.set_column(0, 0, 22)
    h.set_column(1, 1, 42)
    h.set_column(2, 2, 16)
    h.set_column(3, 6, 11)
    h.set_column(7, 7, 40)
    h.set_column(8, 8, 22)
    h.freeze_panes(1, 0)

    total = sum(g['monto'] for _, g in filas) or 1.0
    acumulado, con_nombre = 0.0, 0
    for f, (clave, g) in enumerate(filas, start=1):
        acumulado += g['monto']
        propuesto, origen = proponer(clave)
        if propuesto:
            con_nombre += 1
        h.write(f, 0, clave)
        h.write(f, 1, " / ".join(sorted(g['alias']))[:120])
        h.write(f, 2, g['monto'], pesos)
        h.write(f, 3, g['veces'])
        h.write(f, 4, str(g['primera'] or ""), gris)
        h.write(f, 5, str(g['ultima'] or ""), gris)
        h.write(f, 6, "%.0f%%" % (100.0 * acumulado / total), gris)
        h.write(f, 7, propuesto, llenar)
        h.write(f, 8, origen, gris)
    return total, con_nombre


indice_clientes = {}
for c in P.search([('customer_rank', '>', 0)]):
    indice_clientes.setdefault(" ".join((c.name or "").split()).upper(), c)


def por_rut(rut):
    n = ficha(rut)
    if n:
        return n, "ficha de Odoo"
    n = conocidos.get(rut, "")
    return (n, "planilla de RUT") if n else ("", "")


def por_alias(alias):
    # Solo si el nombre del banco es IGUAL al de un cliente. Parecerse no
    # basta: en esta misma cartola RESTAURANT PAC se parece un 83% a
    # RESTAURANT KEKA y no tienen nada que ver, y dar por pagada la factura de
    # otro no es un dato mal puesto, es un cobro que se deja de perseguir.
    c = indice_clientes.get(" ".join(str(alias or "").split()).upper())
    return (c.name, "nombre igual al cliente") if c else ("", "")


# Los que el banco identifica SOLO por nombre corto van primero: menos filas y
# mas plata.
tot_alias, ya_alias = escribir("Nombres del banco", "Nombre en el banco",
                               orden_alias, por_alias)
tot_rut, ya_rut = escribir("Pagadores por RUT", "RUT", orden, por_rut)
libro.close()


def plata(x):
    return "{:,.0f}".format(x).replace(",", ".")


print("Escrito: %s" % DESTINO)
print()
print("  Hoja 1 - Nombres del banco: %d nombres, %s"
      % (len(orden_alias), plata(tot_alias)))
print("      %d vienen con nombre propuesto" % ya_alias)
print()
print("  Hoja 2 - Pagadores por RUT: %d RUT, %s" % (len(orden), plata(tot_rut)))
print("      %d vienen con nombre propuesto" % ya_rut)
print()
print("Se llena de arriba hacia abajo y se para donde el % acumulado deje de")
print("importar. Despues: tools/importar_pagadores.py")
