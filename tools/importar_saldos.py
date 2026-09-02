"""Carga los saldos de apertura desde la planilla completada.

    odoo-bin shell -c config/odoo.conf -d agrogood_dev --no-http < tools/importar_saldos.py

Por defecto SOLO INFORMA. Para escribir:  AGROGOOD_SALDOS=escribir
La fecha de corte se puede fijar con     AGROGOOD_SALDOS_CORTE=2026-09-15
(si no, es hoy: el saldo cubre hasta el dia en que se carga).

Lee `C:/dev/Saldos por completar.xlsx`, la que genera `exportar_saldos.py`. Se
busca el cliente por su RUT y, si no lo tiene, por el nombre exacto. No se
adivina por parecido: cargarle la deuda al cliente equivocado no da error, da
una llamada a alguien que no debe nada y un cobro que se deja de perseguir.

Volver a cargar la planilla REEMPLAZA el saldo de apertura, no lo suma. Asi
corregir una cifra mal puesta es editar la celda y volver a cargar, y no hace
falta acordarse de cuantas veces se cargo antes.
"""

import os

import openpyxl

from odoo import fields

RUTA = r"C:\dev\Saldos por completar.xlsx"
ESCRIBIR = os.environ.get("AGROGOOD_SALDOS") == "escribir"
CORTE = os.environ.get("AGROGOOD_SALDOS_CORTE") or str(
    fields.Date.context_today(env.user))

print("=" * 74)
print("SALDOS DE APERTURA" + ("  [ESCRIBIENDO]" if ESCRIBIR else "  [SOLO INFORME]"))
print("=" * 74)
print("Fecha de corte: %s" % CORTE)


def normalizar_rut(bruto):
    from odoo.addons.agrogood_crm_reactivation.models.agrogood_payer import (
        normalizar_rut as n)
    return n(bruto)


if not os.path.exists(RUTA):
    print()
    print("No existe %s" % RUTA)
    print("Generala primero con tools/exportar_saldos.py")
else:
    ws = openpyxl.load_workbook(RUTA, data_only=True)["Saldos"]
    Socio = env['res.partner']

    aplicables, sin_encontrar, en_cero, invalidos = [], [], 0, []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila or not (fila[0] or fila[1]):
            continue
        rut = normalizar_rut(fila[0]) if fila[0] else None
        nombre = str(fila[1] or "").strip()
        try:
            saldo = float(fila[5]) if fila[5] not in (None, "") else 0.0
        except (TypeError, ValueError):
            invalidos.append(nombre)
            continue

        socio = Socio.search([('vat', '=', rut)], limit=1) if rut else Socio.browse()
        if not socio:
            socio = Socio.search([('name', '=', nombre),
                                  ('parent_id', '=', False)], limit=1)
        if not socio:
            sin_encontrar.append(nombre or str(fila[0]))
            continue
        if saldo <= 0:
            en_cero += 1
            continue
        aplicables.append((socio, saldo))

    total = sum(s for _, s in aplicables)
    print()
    print("Filas leidas          : %d" % (ws.max_row - 1))
    print("  con saldo que cargar: %d" % len(aplicables))
    print("  en cero (no deben)  : %d" % en_cero)
    print("  cliente no encontrado: %d" % len(sin_encontrar))
    print("  celda ilegible      : %d" % len(invalidos))
    print()
    print("TOTAL POR COBRAR AL ARRANCAR: %s"
          % "{:,.0f}".format(total).replace(",", "."))

    if sin_encontrar:
        print()
        print("-" * 74)
        print("NO SE ENCONTRARON (%d). Se dejan sin cargar." % len(sin_encontrar))
        for n in sin_encontrar[:10]:
            print("   %s" % str(n)[:60])

    if aplicables:
        print()
        print("-" * 74)
        print("LOS DIEZ QUE MAS DEBEN")
        for socio, saldo in sorted(aplicables, key=lambda x: -x[1])[:10]:
            print("   %-44s %14s  %4.1f%%"
                  % (socio.display_name[:44],
                     "{:,.0f}".format(saldo).replace(",", "."),
                     100.0 * saldo / max(total, 1)))

        # Un cero de mas se ve aqui y no en la pantalla de cobranza. No se
        # bloquea -un cliente grande puede deber de verdad mucho- pero se dice,
        # porque el error tipico es teclear 1.500.000 donde iban 150.000 y
        # despues nadie entiende por que ese cliente encabeza la lista.
        mayor, importe_mayor = max(aplicables, key=lambda x: x[1])
        if total and importe_mayor / total > 0.30:
            print()
            print("   OJO: %s concentra el %.0f%% de toda la deuda."
                  % (mayor.display_name, 100.0 * importe_mayor / total))
            print("   Si es de verdad, adelante. Si sobra un cero, se ve aqui.")

    if not ESCRIBIR:
        print()
        print("=" * 74)
        print("SOLO INFORME. Nada se ha modificado.")
        print("Para aplicar: AGROGOOD_SALDOS=escribir")
        print("=" * 74)
    elif aplicables:
        # Se borra lo que hubiera antes: la planilla es la verdad, y cargarla
        # dos veces tiene que dar el mismo resultado que cargarla una.
        anteriores = Socio.search([('agrogood_opening_balance', '>', 0)])
        (anteriores - Socio.browse([s.id for s, _ in aplicables])).write({
            'agrogood_opening_balance': 0.0,
            'agrogood_opening_date': False,
        })
        for socio, saldo in aplicables:
            socio.write({'agrogood_opening_balance': saldo,
                         'agrogood_opening_date': CORTE})
        env.cr.commit()
        con_saldo = Socio.search_count([('agrogood_balance', '>', 0)])
        print()
        print("=" * 74)
        print("CARGADOS: %d saldos de apertura" % len(aplicables))
        print("Clientes que deben algo ahora: %d" % con_saldo)
        print("=" * 74)
        print()
        print("Ya se puede llamar: Paneles > Cobranza, o /agrogood/cobranza en")
        print("el telefono. Y los abonos del banco imputan primero contra este")
        print("saldo, que es la deuda mas antigua que tiene cada cliente.")
