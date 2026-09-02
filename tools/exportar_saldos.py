"""Saca a una planilla la cartera, para anotar cuanto debe cada cliente.

    odoo-bin shell -c config/odoo.conf -d agrogood_dev --no-http < tools/exportar_saldos.py

Escribe `C:/dev/Saldos por completar.xlsx`, fuera del repositorio.

Por que hace falta: Agroapp solo sabe de las entregas que pasan por el sistema,
asi que el dia que arranque la cobranza diria que nadie debe nada. El saldo de
apertura es lo que el cliente ya debia ese dia, de entregas anteriores.

**No es el historico de ventas.** Es una linea por cliente y nada mas. El
detalle entrega por entrega no hace falta para cobrar: al llamar se necesita
saber cuanto debe, no que le vendimos en marzo.

Solo hay una columna que rellenar y viene en cero, porque un saldo no se puede
proponer: hay que mirarlo. Los clientes van agrupados por linea comercial para
poder recorrerlos junto a la contabilidad.
"""

import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SALIDA = r"C:\dev\Saldos por completar.xlsx"

Socio = env['res.partner']
cartera = Socio.search([
    ('agrogood_business_line_id', '!=', False),
    ('parent_id', '=', False),
], order='agrogood_business_line_id, name')

print("=" * 74)
print("SALDOS DE APERTURA POR COMPLETAR")
print("=" * 74)
print("Clientes en la cartera: %d" % len(cartera))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Saldos"

CABECERAS = ["RUT", "Cliente", "Linea comercial", "Telefono",
             "Ultima compra", "SALDO QUE DEBE (completar)"]
ws.append(CABECERAS)
for celda in ws[1]:
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="1C874F")
    celda.alignment = Alignment(horizontal="center")

sin_rut = 0
for s in cartera:
    if not s.vat:
        sin_rut += 1
    ws.append([
        s.vat or "", s.name,
        s.agrogood_business_line_id.name or "",
        s.mobile or s.phone or "",
        s.agrogood_last_order_date or "",
        0,
    ])

anchos = [14, 40, 20, 16, 14, 26]
for i, ancho in enumerate(anchos, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
ws.freeze_panes = "A2"

# La unica columna que se toca, marcada. Una planilla de seis columnas donde
# solo se edita una tiene que decir cual sin que nadie lo explique.
for fila in ws.iter_rows(min_row=2, min_col=6, max_col=6):
    for celda in fila:
        celda.fill = PatternFill("solid", fgColor="FFF3CD")
        celda.number_format = '#,##0'

ws2 = wb.create_sheet("Como se usa")
for linea in [
    "QUE ES ESTO",
    "",
    "Lo que cada cliente debe HOY, de entregas anteriores a Agroapp.",
    "Una sola cifra por cliente, con IVA incluido: lo que le cobrarias si",
    "llamaras ahora mismo.",
    "",
    "QUE NO ES",
    "",
    "No es lo que se le vendio en el ano ni el detalle de sus entregas.",
    "Para cobrar hace falta saber cuanto debe, no que llevo en marzo.",
    "",
    "COMO SE COMPLETA",
    "",
    "1. Se rellena solo la ultima columna, la amarilla.",
    "2. Los que no deben nada se dejan en cero. No hay que borrar filas.",
    "3. Se guarda y se avisa.",
    "",
    "DESPUES",
    "",
    "Ese saldo aparece en la cuenta corriente del cliente y se cobra como",
    "cualquier otra deuda: es lo primero que se paga cuando entra una",
    "transferencia suya, porque es la deuda mas antigua que tiene.",
]:
    ws2.append([linea])
ws2.column_dimensions['A'].width = 76

wb.save(SALIDA)

por_linea = {}
for s in cartera:
    nombre = s.agrogood_business_line_id.name or "(sin linea)"
    por_linea[nombre] = por_linea.get(nombre, 0) + 1

print()
print("%-30s %s" % ("LINEA COMERCIAL", "CLIENTES"))
for nombre, n in sorted(por_linea.items(), key=lambda x: -x[1]):
    print("%-30s %6d" % (nombre, n))
print()
print("Sin RUT en la ficha: %d (se identifican por nombre al cargar)" % sin_rut)
print()
print("Planilla escrita en: %s" % SALIDA)
print("Existe: %s" % os.path.exists(SALIDA))
print()
print("Se rellena solo la ultima columna. Los que no deben nada se dejan en")
print("cero. Despues:")
print("   AGROGOOD_SALDOS=escribir odoo-bin shell ... < tools/importar_saldos.py")
