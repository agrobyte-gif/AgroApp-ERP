import base64
import io
import logging

from odoo import _, fields, models
from odoo.exceptions import UserError

from ..models import cartola

_logger = logging.getLogger(__name__)

# Cuantos abonos se crean de una vez. La cartola de un mes trae mas de trece
# mil filas y crearlas una a una tarda; crearlas todas de golpe se come la
# memoria. En bloques va rapido y no se nota.
BLOQUE = 500


class AgrogoodBankImport(models.TransientModel):
    """Sube el archivo que exporta el banco y lo convierte en abonos.

    Se sube el archivo tal cual sale del banco, sin retocarlo. Pedir que
    alguien lo prepare antes -borrar filas, renombrar columnas- es pedir un
    paso manual todos los meses, y el paso manual es justo lo que se venia a
    quitar.

    El archivo NO se guarda. Lleva RUT, montos y numeros de cuenta de terceros,
    y de el solo hacen falta los abonos: una vez leidos, el original no aporta
    nada que justifique conservarlo.
    """

    _name = 'agrogood.bank.import'
    _description = "Cargar la cartola del banco"

    file = fields.Binary(string="Archivo del banco", required=True)
    filename = fields.Char(string="Nombre del archivo")
    result = fields.Text(string="Resultado", readonly=True)
    movement_ids = fields.Many2many(
        comodel_name='agrogood.bank.movement', string="Abonos cargados",
    )

    def action_cargar(self):
        self.ensure_one()
        try:
            import openpyxl
        except ImportError:
            raise UserError(_(
                "Falta la libreria openpyxl para leer planillas. Se instala "
                "con: pip install openpyxl"))

        nombre = (self.filename or "").lower()
        if not nombre.endswith(('.xlsx', '.xlsm')):
            raise UserError(_(
                "El archivo tiene que ser la planilla que exporta el banco "
                "(.xlsx). Si el banco entrega un PDF, hay que pedirle a la "
                "sucursal el mismo movimiento en Excel."))

        try:
            libro = openpyxl.load_workbook(
                io.BytesIO(base64.b64decode(self.file)),
                read_only=True, data_only=True)
        except Exception as e:
            _logger.warning("No se pudo abrir la cartola: %s", e)
            raise UserError(_(
                "No se pudo abrir el archivo. Suele pasar cuando se guardo "
                "como .xls antiguo: hay que volver a guardarlo como .xlsx."))

        abonos, descartes, hojas_sin_reconocer = [], {}, []
        hojas = list(libro.sheetnames)
        for hoja in hojas:
            filas = libro[hoja].iter_rows(values_only=True)
            leidos, desc = cartola.leer_hoja(filas, nombre_hoja=hoja)
            if desc.get('sin_cabecera'):
                hojas_sin_reconocer.append(hoja)
                continue
            abonos.extend(leidos)
            for k, v in desc.items():
                descartes[k] = descartes.get(k, 0) + v
        libro.close()

        if not abonos:
            raise UserError(_(
                "No se reconocio ningun abono. El lector espera las columnas "
                "que traen Scotiabank ('Rut Origen') y Santander "
                "('CARGO/ABONO'). Hojas leidas: %(hojas)s",
                hojas=", ".join(hojas)))

        creados = self._crear(abonos)
        creados._cruzar()

        self.movement_ids = creados
        self.result = self._resumen(abonos, creados, descartes, hojas_sin_reconocer)
        return {
            'type': 'ir.actions.act_window', 'res_model': self._name,
            'res_id': self.id, 'view_mode': 'form', 'target': 'new',
            'context': dict(self.env.context, agrogood_cargado=True),
        }

    def _crear(self, abonos):
        """Crea los abonos que aun no estaban. Devuelve los creados."""
        Movimiento = self.env['agrogood.bank.movement']
        claves = [a['clave'] for a in abonos]
        ya_estaban = set()
        for i in range(0, len(claves), BLOQUE):
            ya_estaban.update(Movimiento.search([
                ('unique_key', 'in', claves[i:i + BLOQUE]),
            ]).mapped('unique_key'))

        vals = [{
            'bank': a['banco'],
            'sheet': a['hoja'],
            'date': a['fecha'],
            'amount': a['monto'],
            'payer_rut': a['rut'] or False,
            'payer_alias': a['alias'] or False,
            'payer_name': a['nombre_banco'] or False,
            'description': a['descripcion'] or False,
            'source_account': a['cuenta_origen'] or False,
            'unique_key': a['clave'],
        } for a in abonos if a['clave'] not in ya_estaban]

        creados = Movimiento.browse()
        for i in range(0, len(vals), BLOQUE):
            creados |= Movimiento.create(vals[i:i + BLOQUE])
        return creados

    def _resumen(self, abonos, creados, descartes, hojas_sin_reconocer):
        repetidos = len(abonos) - len(creados)
        por_estado = {}
        for m in creados:
            por_estado[m.state] = por_estado.get(m.state, 0) + 1
        identificados = por_estado.get('identified', 0)
        pendientes = por_estado.get('unknown', 0) + por_estado.get('doubtful', 0)

        lineas = [
            _("Abonos en el archivo: %(n)s", n=len(abonos)),
            _("Cargados ahora: %(n)s", n=len(creados)),
        ]
        if repetidos:
            lineas.append(_("Ya estaban de una carga anterior: %(n)s", n=repetidos))
        if descartes.get('cargos'):
            lineas.append(_("Cargos y traspasos, que no los paga ningun "
                            "cliente: %(n)s", n=descartes['cargos']))
        lineas.append("")
        lineas.append(_("Se sabe quien pago: %(n)s", n=identificados))
        lineas.append(_("Falta identificar: %(n)s", n=pendientes))
        if pendientes:
            lineas.append("")
            lineas.append(_(
                "Conviene empezar por los de mas monto. Cada uno que se enlaza "
                "ensena a quien pertenece ese pagador, y con 'Volver a cruzar' "
                "quedan resueltos de golpe todos los demas que vengan de el."))
        if hojas_sin_reconocer:
            lineas.append("")
            lineas.append(_("Hojas que no se reconocieron: %(h)s",
                            h=", ".join(hojas_sin_reconocer)))
        return "\n".join(lineas)

    def action_ver(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _("Abonos cargados"),
            'res_model': 'agrogood.bank.movement',
            'view_mode': 'list,form',
            'domain': [('id', 'in', self.movement_ids.ids)],
            'context': {'search_default_f_pendientes': 1},
        }
